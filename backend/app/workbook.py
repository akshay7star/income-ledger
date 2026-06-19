from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

from .financial_year import financial_year_for
from .repositories import add_expense, add_income_record, dashboard_data, list_documents, list_users
from .tax import calculate_tax_options
from .tax_reconciliation import tax_statement_report


SHEETS = [
    "Summary",
    "Income Records",
    "Expenses",
    "GST",
    "Tax",
    "Tax Documents",
    "Tax Reconciliation",
    "Tax Findings",
    "Documents",
    "Balance Sheet",
    "Import Income",
    "Import Expenses",
]


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel import/export. Install project requirements first.") from exc
    return Workbook, load_workbook


def _append_dict_rows(sheet, rows: list[dict], fields: list[str]) -> None:
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field) for field in fields])


def _selection_values(value: str | list[str] | None, fallback: str = "all") -> list[str]:
    if value is None:
        return [fallback]
    if isinstance(value, str):
        items = [item.strip() for item in value.split(",") if item.strip()]
        return items or [fallback]
    return [str(item).strip() for item in value if str(item).strip()] or [fallback]


def _user_name_map() -> dict[str, str]:
    return {"all": "All users", **{str(user["id"]): user["name"] for user in list_users()}}


def _selected_user_ids(selected_users: list[str]) -> list[str]:
    users = [str(user["id"]) for user in list_users()]
    if "all" in selected_users:
        return users
    return [user_id for user_id in selected_users if user_id in users]


def _append_rows_with_context(sheet, rows: list[dict], fields: list[str], user_name: str, financial_year: str) -> None:
    for row in rows:
        sheet.append([user_name, financial_year, *[row.get(field) for field in fields]])


def _sum_records(records: list[dict], income_type: str | None, field: str) -> float:
    return round(sum(float(row.get(field) or 0) for row in records if income_type is None or row.get("income_type") == income_type), 2)


def _append_monthly_gst_rows(sheet, user_name: str, financial_year: str, monthly: list[dict]) -> None:
    for row in monthly:
        gst_collected = round(float(row.get("gst") or 0), 2)
        gst_input = round(float(row.get("expense_gst") or 0), 2)
        net_gst = round(gst_collected - gst_input, 2)
        if gst_collected == 0 and gst_input == 0 and net_gst == 0:
            continue
        sheet.append(
            [
                user_name,
                financial_year,
                row.get("month"),
                gst_collected,
                gst_input,
                net_gst,
            ]
        )


def _append_accounting_statement(sheet, user_name: str, financial_year: str, data: dict) -> None:
    summary = data["summary"]
    records = data["records"]
    expenses = data["expenses"]
    gross_receipts = float(summary.get("freelance_income") or 0)
    gst_collected = float(summary.get("freelance_gst_collected") or 0)
    gst_input = float(summary.get("expense_gst_claims") or 0)
    total_expenses = float(summary.get("total_expenses") or 0)
    net_expenses = max(0.0, total_expenses - gst_input)
    tds_receivable = _sum_records(records, "freelance_invoice", "tds_amount")
    net_profit = round(gross_receipts - net_expenses, 2)
    cash_bank = round(gross_receipts - tds_receivable + gst_collected - total_expenses, 2)

    start_row = sheet.max_row + 2 if sheet.max_row > 1 else sheet.max_row + 1
    sheet.append([user_name, financial_year, "Profit and Loss", "Particulars", "Debit", "Credit"])
    for category in sorted({expense.get("category") for expense in expenses if expense.get("category")}):
        amount = sum(float(expense.get("amount") or 0) - float(expense.get("gst_amount") or 0) for expense in expenses if expense.get("category") == category)
        sheet.append([user_name, financial_year, "Profit and Loss", f"Expense - {category}", round(amount, 2), 0])
    if not expenses:
        sheet.append([user_name, financial_year, "Profit and Loss", "Business expenses", 0, 0])
    sheet.append([user_name, financial_year, "Profit and Loss", "Net profit transferred to capital", max(0.0, net_profit), max(0.0, -net_profit)])
    sheet.append([user_name, financial_year, "Profit and Loss", "Freelance gross receipts", 0, gross_receipts])
    pnl_debit = round(net_expenses + max(0.0, net_profit), 2)
    pnl_credit = round(gross_receipts + max(0.0, -net_profit), 2)
    sheet.append([user_name, financial_year, "Profit and Loss", "Total", pnl_debit, pnl_credit])
    sheet.append([user_name, financial_year, "Profit and Loss", "Check difference", round(pnl_debit - pnl_credit, 2), ""])

    sheet.append([])
    sheet.append([user_name, financial_year, "Balance Sheet", "Assets / Liabilities", "Debit", "Credit"])
    sheet.append([user_name, financial_year, "Assets", "Cash / bank balance derived from receipts and payments", max(0.0, cash_bank), max(0.0, -cash_bank)])
    sheet.append([user_name, financial_year, "Assets", "TDS receivable from freelance invoices", tds_receivable, 0])
    sheet.append([user_name, financial_year, "Assets", "GST input credit", gst_input, 0])
    sheet.append([user_name, financial_year, "Liabilities", "GST output payable", 0, gst_collected])
    sheet.append([user_name, financial_year, "Capital", "Owner capital / current account", 0 if net_profit >= 0 else abs(net_profit), max(0.0, net_profit)])
    bs_debit = round(max(0.0, cash_bank) + tds_receivable + gst_input + (0 if net_profit >= 0 else abs(net_profit)), 2)
    bs_credit = round(max(0.0, -cash_bank) + gst_collected + max(0.0, net_profit), 2)
    sheet.append([user_name, financial_year, "Balance Sheet", "Total", bs_debit, bs_credit])
    sheet.append([user_name, financial_year, "Balance Sheet", "Check difference", round(bs_debit - bs_credit, 2), ""])

    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=5, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'


def _append_tax_statement_export(tax_doc_sheet, tax_recon_sheet, tax_findings_sheet, user_id: str, user_name: str, financial_year: str) -> None:
    report = tax_statement_report(user_id, financial_year)
    for document in report.get("tax_documents", []):
        tax_doc_sheet.append(
            [
                user_name,
                financial_year,
                document.get("id"),
                document.get("document_id"),
                document.get("document_name"),
                document.get("source_type"),
                document.get("document_status"),
                document.get("is_active"),
                document.get("pan"),
                document.get("tan"),
                document.get("deductor_name"),
                document.get("certificate_number"),
                document.get("period_from"),
                document.get("period_to"),
                document.get("assessment_year"),
                document.get("confidence"),
                document.get("uploaded_at"),
            ]
        )

    for item in report.get("employer_comparisons", []):
        tax_recon_sheet.append(
            [
                user_name,
                financial_year,
                "Salary",
                item.get("employer"),
                item.get("tan"),
                "",
                "192",
                item.get("ledger_salary"),
                item.get("form16_salary"),
                item.get("form26as_amount"),
                item.get("ledger_tds"),
                item.get("form16_tds"),
                item.get("form26as_tds"),
                item.get("salary_difference"),
                item.get("tds_difference"),
                item.get("status"),
            ]
        )
    for item in report.get("monthly_salary_comparisons", []):
        tax_recon_sheet.append(
            [
                user_name,
                financial_year,
                "Salary Month",
                item.get("employer"),
                item.get("tan"),
                item.get("month"),
                "192",
                item.get("ledger_salary"),
                "",
                item.get("form26as_amount"),
                item.get("ledger_tds"),
                "",
                item.get("form26as_tds"),
                item.get("amount_difference"),
                item.get("tds_difference"),
                item.get("status"),
            ]
        )
    for item in report.get("freelance_comparisons", []):
        tax_recon_sheet.append(
            [
                user_name,
                financial_year,
                "Freelance",
                item.get("deductor_name"),
                item.get("tan"),
                "",
                ", ".join(item.get("sections") or []),
                item.get("ledger_receipts"),
                "",
                item.get("form26as_amount"),
                item.get("ledger_tds"),
                "",
                item.get("form26as_tds"),
                item.get("receipt_difference"),
                item.get("tds_difference"),
                item.get("status"),
            ]
        )

    for finding in report.get("findings", []):
        tax_findings_sheet.append(
            [
                user_name,
                financial_year,
                finding.get("severity"),
                finding.get("type"),
                finding.get("message"),
                finding.get("deductor_name"),
                finding.get("tan"),
                finding.get("certificate_number"),
                finding.get("record_id"),
                finding.get("tax_document_id"),
                finding.get("month"),
                finding.get("ledger_salary"),
                finding.get("form26as_amount"),
                finding.get("ledger_tds"),
                finding.get("form26as_tds"),
                finding.get("amount_difference"),
                finding.get("tds_difference"),
                finding.get("transaction_date"),
            ]
        )


def _format_workbook(wb) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F2937")
    section_fill = PatternFill("solid", fgColor="E5E7EB")
    check_fill = PatternFill("solid", fgColor="DCFCE7")
    header_font = Font(bold=True, color="FFFFFF")
    section_font = Font(bold=True, color="111827")

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for column_cells in sheet.columns:
                letter = get_column_letter(column_cells[0].column)
                width = min(42, max(12, max(len(str(cell.value or "")) for cell in column_cells[:100]) + 2))
                sheet.column_dimensions[letter].width = width

    balance = wb["Balance Sheet"]
    for row in balance.iter_rows(min_row=2, max_row=balance.max_row):
        label = row[3].value if len(row) >= 4 else None
        if label in {"Particulars", "Assets / Liabilities", "Total"}:
            for cell in row:
                cell.fill = section_fill
                cell.font = section_font
        if label == "Check difference":
            for cell in row:
                cell.fill = check_fill
                cell.font = section_font


def create_workbook_export(
    user_id: str | None = None,
    financial_year: str | None = None,
    user_ids: str | list[str] | None = None,
    financial_years: str | list[str] | None = None,
) -> Path:
    Workbook, _load_workbook = _load_openpyxl()
    selected_users = _selection_values(user_ids if user_ids is not None else user_id, "all")
    selected_users = _selected_user_ids(selected_users)
    selected_years = _selection_values(financial_years if financial_years is not None else financial_year, "")
    selected_years = [year for year in selected_years if year]
    if not selected_years:
        raise ValueError("At least one financial year is required.")
    if not selected_users:
        raise ValueError("At least one valid user is required.")
    users = _user_name_map()
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEETS:
        wb.create_sheet(name)

    summary = wb["Summary"]
    summary.append(["User", "Financial Year", "Metric", "Value"])
    income_sheet = wb["Income Records"]
    income_fields = ["id", "user_id", "record_date", "period_label", "income_type", "payer", "gross_amount", "net_amount", "tds_amount", "deductions_amount", "pf_amount", "vpf_amount", "gst_amount"]
    income_sheet.append(["User", "Financial Year", *income_fields])
    expense_sheet = wb["Expenses"]
    expense_fields = ["id", "user_id", "expense_date", "category", "amount", "gst_amount", "notes"]
    expense_sheet.append(["User", "Financial Year", *expense_fields])
    gst_sheet = wb["GST"]
    gst_sheet.append(["User", "Financial Year", "Month", "GST Collected", "GST Input Claims", "Net GST Payable"])
    tax_sheet = wb["Tax"]
    tax_sheet.append(["User", "Financial Year", "Metric", "Value"])
    tax_doc_sheet = wb["Tax Documents"]
    tax_doc_sheet.append([
        "User",
        "Financial Year",
        "Tax Document ID",
        "Document ID",
        "Document Name",
        "Source Type",
        "Document Status",
        "Active",
        "PAN",
        "TAN",
        "Deductor",
        "Certificate Number",
        "Period From",
        "Period To",
        "Assessment Year",
        "Confidence",
        "Uploaded At",
    ])
    tax_recon_sheet = wb["Tax Reconciliation"]
    tax_recon_sheet.append([
        "User",
        "Financial Year",
        "Area",
        "Party",
        "TAN",
        "Month",
        "Section",
        "Ledger Amount",
        "Form 16 Amount",
        "26AS Amount",
        "Ledger TDS",
        "Form 16 TDS",
        "26AS TDS",
        "Amount Difference",
        "TDS Difference",
        "Status",
    ])
    tax_findings_sheet = wb["Tax Findings"]
    tax_findings_sheet.append([
        "User",
        "Financial Year",
        "Severity",
        "Type",
        "Message",
        "Deductor",
        "TAN",
        "Certificate Number",
        "Record ID",
        "Tax Document ID",
        "Month",
        "Ledger Salary",
        "26AS Amount",
        "Ledger TDS",
        "26AS TDS",
        "Amount Difference",
        "TDS Difference",
        "Transaction Date",
    ])
    balance_sheet = wb["Balance Sheet"]
    balance_sheet.append(["User", "Financial Year", "Statement", "Particulars", "Debit", "Credit"])

    for selected_user in selected_users:
        user_name = users.get(str(selected_user), f"User {selected_user}")
        for selected_year in selected_years:
            data = dashboard_data(selected_user, selected_year)
            for key, value in data["summary"].items():
                summary.append([user_name, selected_year, key, value])

            _append_rows_with_context(income_sheet, data["records"], income_fields, user_name, selected_year)
            _append_rows_with_context(expense_sheet, data["expenses"], expense_fields, user_name, selected_year)

            _append_monthly_gst_rows(gst_sheet, user_name, selected_year, data["monthly"])

            try:
                tax = calculate_tax_options(selected_year, data["summary"].get("salary_income", 0), data["summary"].get("freelance_profit", 0))["selected"]
                for key in ["regime", "taxable_income", "base_tax", "rebate", "marginal_relief", "cess", "total_tax"]:
                    tax_sheet.append([user_name, selected_year, key, tax.get(key)])
                tax_sheet.append([user_name, selected_year, "tds_paid", data["summary"].get("tds_paid", 0)])
                tax_sheet.append([user_name, selected_year, "remaining_tax", max(0, tax.get("total_tax", 0) - data["summary"].get("tds_paid", 0))])
            except KeyError as exc:
                tax_sheet.append([user_name, selected_year, "tax_status", f"Unsupported tax year: {exc}"])

            _append_accounting_statement(balance_sheet, user_name, selected_year, data)
            _append_tax_statement_export(tax_doc_sheet, tax_recon_sheet, tax_findings_sheet, selected_user, user_name, selected_year)

    documents = list_documents()
    document_sheet = wb["Documents"]
    document_fields = ["id", "original_name", "document_type", "status", "detected_user_id", "confidence", "uploaded_at"]
    document_sheet.append(["User", "Financial Year", *document_fields])
    for doc in documents:
        detected_user = str(doc.get("detected_user_id") or "all")
        extracted = doc.get("extracted") or {}
        extracted_year = extracted.get("financial_year") or ""
        if not extracted_year and extracted.get("record_date"):
            try:
                extracted_year = financial_year_for(extracted.get("record_date"))
            except ValueError:
                extracted_year = ""
        if detected_user not in selected_users:
            continue
        if extracted_year and extracted_year not in selected_years:
            continue
        document_sheet.append([users.get(detected_user, detected_user), extracted_year, *[doc.get(field) for field in document_fields]])

    wb["Import Income"].append(["user_id", "income_type", "record_date", "payer", "gross_amount", "net_amount", "tds_amount", "deductions_amount", "pf_amount", "vpf_amount", "gst_amount"])
    wb["Import Expenses"].append(["user_id", "expense_date", "category", "amount", "gst_amount", "notes"])

    label = "multi-year" if len(selected_years) > 1 else selected_years[0].replace(" ", "-").replace("/", "-")
    path = Path(tempfile.gettempdir()) / f"income-ledger-{label}.xlsx"
    _format_workbook(wb)
    wb.save(path)
    return path


def create_import_template() -> Path:
    Workbook, _load_workbook = _load_openpyxl()
    wb = Workbook()
    wb.active.title = "Import Income"
    wb["Import Income"].append(["user_id", "income_type", "record_date", "payer", "gross_amount", "net_amount", "tds_amount", "deductions_amount", "pf_amount", "vpf_amount", "gst_amount"])
    expense_sheet = wb.create_sheet("Import Expenses")
    expense_sheet.append(["user_id", "expense_date", "category", "amount", "gst_amount", "notes"])
    path = Path(tempfile.gettempdir()) / "income-ledger-import-template.xlsx"
    wb.save(path)
    return path


def _row_dict(headers: list[Any], row: tuple[Any, ...]) -> dict:
    return {str(header).strip(): value for header, value in zip(headers, row) if header}


def import_workbook(path: Path) -> dict:
    _Workbook, load_workbook = _load_openpyxl()
    wb = load_workbook(path, data_only=True)
    created = {"income": 0, "expenses": 0}
    errors: list[dict] = []

    if "Import Income" in wb.sheetnames:
        sheet = wb["Import Income"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            payload = _row_dict(headers, row)
            try:
                add_income_record(payload)
                created["income"] += 1
            except Exception as exc:
                errors.append({"sheet": "Import Income", "row": index, "error": str(exc)})

    if "Import Expenses" in wb.sheetnames:
        sheet = wb["Import Expenses"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            payload = _row_dict(headers, row)
            try:
                add_expense(payload)
                created["expenses"] += 1
            except Exception as exc:
                errors.append({"sheet": "Import Expenses", "row": index, "error": str(exc)})

    return {"created": created, "errors": errors}
