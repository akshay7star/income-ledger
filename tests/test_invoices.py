from __future__ import annotations

import json
from pathlib import Path

import pytest


def test_existing_fixed_tds_invoice_is_migrated_to_percentage(tmp_path, monkeypatch):
    from backend.app import database

    data_dir = tmp_path / "migration-data"
    monkeypatch.setattr(database, "DATA_DIR", data_dir)
    monkeypatch.setattr(database, "UPLOAD_DIR", data_dir / "uploads")
    monkeypatch.setattr(database, "GENERATED_INVOICE_DIR", data_dir / "generated_invoices")
    monkeypatch.setattr(database, "DB_PATH", data_dir / "income_ledger.sqlite3")
    database.init_db()

    with database.get_connection() as conn:
        seller_id = conn.execute(
            "INSERT INTO invoice_profiles (display_name) VALUES ('Legacy seller')"
        ).lastrowid
        client_id = conn.execute(
            "INSERT INTO invoice_clients (client_name) VALUES ('Legacy client')"
        ).lastrowid
        conn.execute(
            """
            INSERT INTO generated_invoices (
                invoice_number, invoice_date, financial_year, seller_profile_id,
                client_id, subtotal_amount, tds_rate, tds_amount
            ) VALUES ('LEGACY-1', '2026-07-01', 'FY 2026-27', ?, ?, 2000, 0, 200)
            """,
            (seller_id, client_id),
        )
        conn.execute("ALTER TABLE generated_invoices DROP COLUMN tds_rate")

    database.init_db()

    with database.get_connection() as conn:
        migrated = conn.execute(
            "SELECT tds_rate, tds_amount FROM generated_invoices WHERE invoice_number = 'LEGACY-1'"
        ).fetchone()
    assert migrated["tds_rate"] == 10
    assert migrated["tds_amount"] == 200


@pytest.fixture
def invoice_env(tmp_path, monkeypatch):
    pytest.importorskip("reportlab")
    from backend.app import database, invoices
    from backend.app.repositories import create_user

    data_dir = tmp_path / "data"
    generated_dir = data_dir / "generated_invoices"
    monkeypatch.setattr(database, "DATA_DIR", data_dir)
    monkeypatch.setattr(database, "UPLOAD_DIR", data_dir / "uploads")
    monkeypatch.setattr(database, "GENERATED_INVOICE_DIR", generated_dir)
    monkeypatch.setattr(database, "DB_PATH", data_dir / "income_ledger.sqlite3")
    monkeypatch.setattr(invoices, "GENERATED_INVOICE_DIR", generated_dir)
    database.init_db()

    user = create_user({"name": "Ledger User", "pan": "", "aliases": "", "profile_hints": ""})
    seller = invoices.create_invoice_profile(
        {
            "display_name": "Example Consultant",
            "legal_name": "Example Consultant",
            "address_line1": "1 Seller Street",
            "city": "Noida",
            "state_name": "Uttar Pradesh",
            "state_code": "09",
            "postal_code": "201301",
            "email": "seller@example.com",
            "gstin": "09ABCDE1234F1Z5",
            "pan": "ABCDE1234F",
            "signature_label": "Authorised Signatory",
            "is_default": True,
        }
    )
    client = invoices.create_invoice_client(
        {
            "client_name": "Example Client Private Limited",
            "legal_name": "Example Client Private Limited",
            "address_line1": "2 Client Avenue",
            "city": "Noida",
            "state_name": "Uttar Pradesh",
            "state_code": "09",
            "postal_code": "201301",
            "email": "client@example.com",
            "gstin": "09AAAAA0000A1Z5",
            "pan": "AAAAA0000A",
        }
    )
    return {
        "database": database,
        "invoices": invoices,
        "data_dir": data_dir,
        "generated_dir": generated_dir,
        "user": user,
        "seller": seller,
        "client": client,
    }


def invoice_payload(invoice_env, **overrides):
    payload = {
        "invoice_number": "004/2026-27",
        "invoice_date": "2026-07-27",
        "billable_period": "July 2026",
        "seller_profile_id": invoice_env["seller"]["id"],
        "client_id": invoice_env["client"]["id"],
        "ledger_user_id": None,
        "place_of_supply": "Uttar Pradesh",
        "gst_treatment": "auto",
        "payment_terms": "Due on receipt",
        "ship_to_same_as_bill_to": True,
        "tds_rate": 10,
        "notes": "Thank you.",
        "items": [
            {
                "description": "Professional Charges for the M/o July, 2026",
                "hsn_sac": "998314",
                "quantity": 1,
                "unit": "Nos",
                "rate": 18,
                "amount": 283333,
            }
        ],
    }
    payload.update(overrides)
    return payload


def test_invoice_calculations_same_state_inter_state_and_no_gst(invoice_env):
    invoices = invoice_env["invoices"]
    base = invoice_payload(invoice_env)

    same_state = invoices.calculate_invoice(base, invoice_env["seller"], invoice_env["client"])
    assert same_state["gst_treatment"] == "same_state"
    assert same_state["subtotal_amount"] == 283333
    assert same_state["cgst_amount"] == 25499.97
    assert same_state["sgst_amount"] == 25499.97
    assert same_state["igst_amount"] == 0
    assert same_state["grand_total_amount"] == 334332.94
    assert same_state["tds_rate"] == 10
    assert same_state["tds_amount"] == 28333.30
    assert same_state["net_receivable_amount"] == 305999.64
    assert same_state["items"][0]["rate"] == 18
    assert same_state["items"][0]["amount"] == 283333
    assert same_state["amount_words"].startswith("INR Three Lakh Thirty Four Thousand")

    other_state_client = {**invoice_env["client"], "state_code": "07", "state_name": "Delhi"}
    inter_state = invoices.calculate_invoice(base, invoice_env["seller"], other_state_client)
    assert inter_state["gst_treatment"] == "inter_state"
    assert inter_state["cgst_amount"] == 0
    assert inter_state["sgst_amount"] == 0
    assert inter_state["igst_amount"] == 50999.94

    no_gst = invoices.calculate_invoice(
        {**base, "gst_treatment": "no_gst"}, invoice_env["seller"], invoice_env["client"]
    )
    assert no_gst["gst_amount"] == 0
    assert no_gst["grand_total_amount"] == 283333


def test_legacy_unit_rate_and_fixed_tds_payload_is_normalized(invoice_env):
    legacy = invoice_payload(invoice_env)
    legacy.pop("tds_rate")
    legacy["tds_amount"] = 200
    legacy["items"] = [
        {
            "description": "Legacy consulting line",
            "quantity": 2,
            "unit": "Hours",
            "rate": 1000,
            "gst_rate": 18,
        }
    ]

    calculated = invoice_env["invoices"].calculate_invoice(
        legacy, invoice_env["seller"], invoice_env["client"]
    )

    assert calculated["subtotal_amount"] == 2000
    assert calculated["items"][0]["amount"] == 2000
    assert calculated["items"][0]["rate"] == 18
    assert calculated["tds_rate"] == 10
    assert calculated["tds_amount"] == 200


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("items", [], "At least one"),
        (
            "items",
            [{"description": "Service", "quantity": 0, "rate": 18, "amount": 100}],
            "quantity must be greater",
        ),
        ("tds_rate", 101, "between 0% and 100%"),
    ],
)
def test_invoice_validation(invoice_env, field, value, message):
    with pytest.raises(ValueError, match=message):
        invoice_env["invoices"].create_invoice_draft(
            invoice_payload(invoice_env, **{field: value})
        )


def test_draft_crud_uniqueness_and_profile_reference_rules(invoice_env):
    invoices = invoice_env["invoices"]
    draft = invoices.create_invoice_draft(
        invoice_payload(invoice_env, invoice_date="2027-03-15")
    )
    assert draft["status"] == "draft"
    assert draft["financial_year"] == "FY 2026-27"
    assert len(draft["items"]) == 1

    with pytest.raises(invoices.InvoiceConflictError, match="already exists"):
        invoices.create_invoice_draft(invoice_payload(invoice_env))

    updated = invoices.update_invoice_draft(
        draft["id"],
        invoice_payload(
            invoice_env,
            invoice_number="005/2026-27",
            items=[
                {
                    "description": "Consulting",
                    "hsn_sac": "998314",
                    "quantity": 2,
                    "unit": "Hours",
                    "rate": 18,
                    "amount": 2000,
                }
            ],
            tds_rate=0,
        ),
    )
    assert updated["invoice_number"] == "005/2026-27"
    assert updated["subtotal_amount"] == 2000
    assert updated["items"][0]["line_number"] == 1
    assert updated["items"][0]["rate"] == 18
    assert updated["items"][0]["amount"] == 2000

    with pytest.raises(invoices.InvoiceConflictError):
        invoices.delete_invoice_profile(invoice_env["seller"]["id"])
    with pytest.raises(invoices.InvoiceConflictError):
        invoices.delete_invoice_client(invoice_env["client"]["id"])

    assert invoices.delete_invoice_draft(draft["id"])["deleted"] is True
    with pytest.raises(KeyError):
        invoices.get_invoice(draft["id"])


def test_preview_issue_link_and_cancel(invoice_env):
    from pypdf import PdfReader

    invoices = invoice_env["invoices"]
    draft = invoices.create_invoice_draft(invoice_payload(invoice_env))
    preview_path = invoices.preview_invoice(draft["id"])
    assert preview_path.exists()
    preview_reader = PdfReader(str(preview_path))
    assert float(preview_reader.pages[0].mediabox.width) == 612
    assert float(preview_reader.pages[0].mediabox.height) == 792
    preview_text = "\n".join(page.extract_text() or "" for page in preview_reader.pages)
    assert "Tax Invoice" in preview_text
    assert "DRAFT" in preview_text
    assert "Professional Charges" in preview_text
    assert "Output-CGST-9%" in preview_text
    assert "Output-SGST-9%" in preview_text
    assert "25,499.97" in preview_text

    issued = invoices.issue_invoice(
        draft["id"],
        create_income_record=True,
        ledger_user_id=invoice_env["user"]["id"],
    )
    assert issued["status"] == "issued"
    assert issued["income_record_id"]
    assert issued["pdf_available"] is True
    final_path = invoices.invoice_pdf_file(draft["id"])
    final_text = "\n".join(page.extract_text() or "" for page in PdfReader(str(final_path)).pages)
    assert "Tax Invoice" in final_text
    assert "DRAFT" not in final_text

    with invoice_env["database"].get_connection() as conn:
        record = conn.execute(
            "SELECT * FROM income_records WHERE id = ?", (issued["income_record_id"],)
        ).fetchone()
        assert record["gross_amount"] == 283333
        assert record["net_amount"] == 254999.7
        metadata = json.loads(record["metadata_json"])
        assert metadata["generated_invoice_id"] == draft["id"]
        assert metadata["gst_amount"] == 50999.94
        assert metadata["tds_rate"] == 10

    with pytest.raises(invoices.InvoiceConflictError):
        invoices.update_invoice_draft(draft["id"], invoice_payload(invoice_env))
    with pytest.raises(invoices.InvoiceConflictError):
        invoices.delete_invoice_draft(draft["id"])

    cancelled = invoices.cancel_invoice(draft["id"], "Incorrect buyer reference")
    assert cancelled["status"] == "cancelled"
    assert cancelled["income_record_id"] == issued["income_record_id"]
    assert cancelled["metadata"]["cancellation"]["reason"] == "Incorrect buyer reference"
    assert final_path.exists()


def test_issue_without_ledger_link_creates_no_income(invoice_env):
    invoices = invoice_env["invoices"]
    draft = invoices.create_invoice_draft(
        invoice_payload(invoice_env, invoice_number="006/2026-27", tds_rate=0)
    )
    issued = invoices.issue_invoice(draft["id"])
    assert issued["status"] == "issued"
    assert issued["income_record_id"] is None
    with invoice_env["database"].get_connection() as conn:
        count = conn.execute("SELECT COUNT(*) AS count FROM income_records").fetchone()["count"]
    assert count == 0


def test_backup_restore_preserves_generated_invoice_pdf(invoice_env, monkeypatch):
    from backend.app import backup

    invoices = invoice_env["invoices"]
    draft = invoices.create_invoice_draft(
        invoice_payload(invoice_env, invoice_number="007/2026-27", tds_rate=0)
    )
    issued = invoices.issue_invoice(draft["id"])
    pdf_path = invoices.invoice_pdf_file(issued["id"])
    backup_result = backup.create_backup()
    pdf_path.unlink()
    assert not pdf_path.exists()

    restored = backup.restore_backup(Path(backup_result["path"]))
    assert restored["restored"] is True
    assert pdf_path.exists()


def test_workbook_export_contains_generated_invoice_sheet(invoice_env):
    from openpyxl import load_workbook
    from backend.app.workbook import create_workbook_export

    invoice_env["invoices"].create_invoice_draft(
        invoice_payload(invoice_env, invoice_number="008/2026-27", tds_rate=0)
    )
    path = create_workbook_export(
        user_ids=[str(invoice_env["user"]["id"])],
        financial_years=["FY 2026-27"],
    )
    workbook = load_workbook(path, data_only=True)
    assert "Generated Invoices" in workbook.sheetnames
    rows = list(workbook["Generated Invoices"].iter_rows(values_only=True))
    assert rows[0][3] == "Invoice Number"
    assert "TDS Rate %" in rows[0]
    assert any(row[3] == "008/2026-27" for row in rows[1:])


def test_invoice_api_routes_are_authenticated_and_cover_lifecycle(invoice_env):
    pytest.importorskip("python_multipart")
    from fastapi.testclient import TestClient
    from backend.app import auth
    from backend.app.main import app

    auth._sessions.clear()
    auth.setup_pin("1234")
    client = TestClient(app)
    assert client.get("/api/invoices").status_code == 401
    token = auth.login("1234")
    headers = {"X-Income-Ledger-Token": token}

    assert client.get("/api/invoice-profiles", headers=headers).status_code == 200
    assert client.get("/api/invoice-clients", headers=headers).status_code == 200
    created = client.post(
        "/api/invoices",
        headers=headers,
        json=invoice_payload(
            invoice_env,
            invoice_number="009/2026-27",
            tds_rate=0,
        ),
    )
    assert created.status_code == 200, created.text
    invoice = created.json()
    assert client.get(f"/api/invoices/{invoice['id']}", headers=headers).status_code == 200

    preview = client.post(f"/api/invoices/{invoice['id']}/preview", headers=headers)
    assert preview.status_code == 200
    assert preview.headers["content-type"] == "application/pdf"

    issued = client.post(
        f"/api/invoices/{invoice['id']}/issue",
        headers=headers,
        json={"create_income_record": False},
    )
    assert issued.status_code == 200, issued.text
    assert issued.json()["status"] == "issued"
    assert client.get(f"/api/invoices/{invoice['id']}/pdf", headers=headers).status_code == 200
    assert client.put(
        f"/api/invoices/{invoice['id']}",
        headers=headers,
        json=invoice_payload(invoice_env, invoice_number="010/2026-27", tds_rate=0),
    ).status_code == 409

    cancelled = client.post(
        f"/api/invoices/{invoice['id']}/cancel",
        headers=headers,
        json={"reason": "API lifecycle test"},
    )
    assert cancelled.status_code == 200
    assert cancelled.json()["status"] == "cancelled"
