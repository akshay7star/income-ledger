import pytest


def setup_workbook_db(tmp_path, monkeypatch):
    from backend.app import database

    data_dir = tmp_path / "data"
    upload_dir = data_dir / "uploads"
    monkeypatch.setattr(database, "DATA_DIR", data_dir)
    monkeypatch.setattr(database, "UPLOAD_DIR", upload_dir)
    monkeypatch.setattr(database, "DB_PATH", data_dir / "income_ledger.sqlite3")
    database.init_db()
    upload_dir.mkdir(parents=True, exist_ok=True)
    return database


def test_import_template_contains_import_sheets():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    from backend.app.workbook import create_import_template

    path = create_import_template()
    wb = load_workbook(path)

    assert "Import Income" in wb.sheetnames
    assert "Import Expenses" in wb.sheetnames


def test_workbook_import_creates_manual_rows_and_reports_errors(tmp_path, monkeypatch):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    from backend.app.workbook import import_workbook

    database = setup_workbook_db(tmp_path, monkeypatch)
    with database.get_connection() as conn:
        conn.execute("INSERT INTO users (id, name, aliases, profile_hints) VALUES (1, 'User', '', '')")

    workbook_path = tmp_path / "import.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Income"
    ws.append(["user_id", "income_type", "record_date", "payer", "gross_amount", "net_amount", "tds_amount"])
    ws.append([1, "salary", "2026-04-01", "Employer", 100, 90, 10])
    ws.append([1, "salary", "", "Bad", 100, 90, 10])
    expenses = wb.create_sheet("Import Expenses")
    expenses.append(["user_id", "expense_date", "category", "amount", "gst_amount", "notes"])
    expenses.append([1, "2026-04-02", "Software", 1000, 180, "IDE"])
    wb.save(workbook_path)

    result = import_workbook(workbook_path)

    assert result["created"]["income"] == 1
    assert result["created"]["expenses"] == 1
    assert result["errors"][0]["sheet"] == "Import Income"


def test_workbook_export_supports_multi_year_context_and_balance_sheet(tmp_path, monkeypatch):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    from backend.app.workbook import create_workbook_export

    database = setup_workbook_db(tmp_path, monkeypatch)
    with database.get_connection() as conn:
        conn.execute("INSERT INTO users (id, name, aliases, profile_hints) VALUES (1, 'User One', '', '')")
        conn.execute("INSERT INTO users (id, name, aliases, profile_hints) VALUES (2, 'User Two', '', '')")
        conn.execute(
            """
            INSERT INTO income_records
                (user_id, financial_year, record_date, period_label, income_type, payer,
                 gross_amount, net_amount, tds_amount, metadata_json)
            VALUES (1, 'FY 2025-26', '2025-04-30', 'Apr 2025', 'freelance_invoice', 'Client',
                    100000, 90000, 10000, '{"gst_amount": 18000}')
            """
        )
        conn.execute(
            """
            INSERT INTO freelance_expenses
                (user_id, financial_year, expense_date, category, amount, gst_amount, notes)
            VALUES (1, 'FY 2025-26', '2025-04-10', 'Software', 20000, 3600, 'IDE')
            """
        )
        conn.execute(
            """
            INSERT INTO income_records
                (user_id, financial_year, record_date, period_label, income_type, payer,
                 gross_amount, net_amount, tds_amount, metadata_json)
            VALUES (1, 'FY 2025-26', '2025-05-30', 'May 2025', 'salary', 'Employer',
                    100000, 90000, 10000, '{"gst_amount": 0}')
            """
        )

    path = create_workbook_export(user_ids="1", financial_years="FY 2025-26,FY 2026-27")
    wb = load_workbook(path, data_only=True)

    assert "Balance Sheet" in wb.sheetnames
    assert wb["Income Records"]["A1"].value == "User"
    assert wb["Income Records"]["B1"].value == "Financial Year"
    assert wb["Income Records"]["A2"].value == "User One"
    assert wb["Income Records"]["B2"].value == "FY 2025-26"
    gst_rows = list(wb["GST"].iter_rows(min_row=2, values_only=True))
    assert all(any(float(value or 0) for value in row[3:6]) for row in gst_rows)
    assert not any(row[2] == "May 2025" for row in gst_rows)
    balance_rows = list(wb["Balance Sheet"].iter_rows(values_only=True))
    assert any(row[3] == "Net profit transferred to capital" and row[4] == 83600 for row in balance_rows)
    assert any(row[2] == "Profit and Loss" and row[3] == "Expense - Software" and row[4] == 16400 for row in balance_rows)
    assert any(row[3] == "Check difference" and row[4] == 0 for row in balance_rows)
